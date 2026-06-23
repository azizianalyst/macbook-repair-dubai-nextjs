#!/usr/bin/env python3
"""One-off seeder: read the MacBook Repair Dubai backlink workbook (tabs 04 Guest Post
Tracker + 10 Web 2.0/Profiles) and write the SEO command-center JSON store.

The store format matches src/lib/seo-store.ts: { "backlinks": [...], "snapshots": [] }.
Run locally so `npm run dev` shows the data; on production, prefer the in-app CSV import
or point SEO_DB at the persistent path and run this against it.

Usage:
  python3 scripts/import-backlinks.py \
    [--xlsx ~/Desktop/websites/MacBookRepairDubai-Backlinks-GuestPost-Master.xlsx] \
    [--out ./data/seo.json]        # default: $SEO_DB or ./data/seo.json
"""
import argparse, json, os, sys, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required: pip3 install openpyxl")

REPO = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path.home() / "Desktop/websites/MacBookRepairDubai-Backlinks-GuestPost-Master.xlsx"
NOW = datetime.datetime.now(datetime.timezone.utc).isoformat()

STATUS_MAP = {
    "qualified": "Prospect", "researching": "Prospect", "to do": "Prospect", "prospect": "Prospect",
    "outreach sent": "Outreach", "outreach": "Outreach", "replied": "Outreach", "negotiating": "Outreach",
    "agreed": "Outreach", "writing": "Drafted", "drafted": "Drafted", "submitted": "Drafted",
    "live": "Live", "indexed": "Live", "owned": "Live", "owned ✓": "Live",
    "rejected": "Rejected", "toxic-skip": "Rejected", "lost": "Lost", "watch": "Watch", "disavow": "Disavow",
}

def cell(row, i):
    v = row[i] if i < len(row) else None
    return "" if v is None else str(v).strip()

def map_status(s, draft=""):
    if draft and "draft" in draft.lower():
        return "Drafted"
    key = (s or "").strip().lower()
    for k, v in STATUS_MAP.items():
        if key.startswith(k):
            return v
    return "Prospect"

def tab04_rows(ws):
    out, started = [], False
    for r in ws.iter_rows(values_only=True):
        c = cell(r, 2)  # col C = site
        if c == "Target site / blog":
            started = True; continue
        if not started or not c:
            continue
        out.append({
            "kind": "prospect", "status": map_status(cell(r, 10), cell(r, 18)),
            "site": c, "source_url": cell(r, 3), "target_url": cell(r, 17),
            "anchor_text": cell(r, 15), "anchor_type": cell(r, 16) if cell(r, 16) in ANCHORS else "",
            "category": "guest", "dr": cell(r, 5), "country": cell(r, 7),
            "contact_email": cell(r, 9) if "@" in cell(r, 9) else "",
            "notes": " · ".join(x for x in [cell(r, 4), cell(r, 23)] if x),  # niche + notes
        })
    return out

TYPE_CAT = {"citation": "citation", "web 2.0": "community", "profile": "community",
            "haro": "haro", "q&a": "community", "social": "community"}

def tab10_rows(ws):
    out, started = [], False
    for r in ws.iter_rows(values_only=True):
        p = cell(r, 1)  # col B = platform
        if p == "Platform":
            started = True; continue
        if not started or not p:
            continue
        typ = cell(r, 2).lower()
        out.append({
            "kind": "prospect", "status": map_status(cell(r, 6)),
            "site": p, "source_url": cell(r, 4),
            "target_url": "https://macbook-repair-dubai.ae/",
            "category": TYPE_CAT.get(typ, "directory"),
            "notes": " · ".join(x for x in [cell(r, 2), cell(r, 3)] if x),  # type + what-to-do
        })
    return out

ANCHORS = {"Branded", "Naked URL", "Partial-match", "Exact-match", "Generic", "Other"}

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    ap.add_argument("--out", default=os.environ.get("SEO_DB") or str(REPO / "data" / "seo.json"))
    args = ap.parse_args()

    if not Path(args.xlsx).exists():
        sys.exit(f"workbook not found: {args.xlsx}")
    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    rows = []
    for name in wb.sheetnames:
        if name.startswith("04"): rows += tab04_rows(wb[name])
        if name.startswith("10"): rows += tab10_rows(wb[name])

    out = Path(args.out)
    db = {"backlinks": [], "snapshots": []}
    if out.exists():
        try: db = json.loads(out.read_text()) or db
        except Exception: pass
    db.setdefault("backlinks", []); db.setdefault("snapshots", [])

    seen = {f"{b.get('site','')}|{b.get('target_url','')}".lower() for b in db["backlinks"]}
    next_id = max([b.get("id", 0) for b in db["backlinks"]], default=0) + 1
    added = 0
    for r in rows:
        key = f"{r['site']}|{r.get('target_url','')}".lower()
        if key in seen: continue
        seen.add(key)
        r.update({"id": next_id, "created_at": NOW, "updated_at": NOW, "first_seen": NOW, "last_checked": None})
        db["backlinks"].append(r); next_id += 1; added += 1

    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(db, indent=2))
    print(f"Imported {added} backlinks (total {len(db['backlinks'])}) -> {out}")

if __name__ == "__main__":
    main()
